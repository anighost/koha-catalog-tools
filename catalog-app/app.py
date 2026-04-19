"""
app.py — Dishari Catalog Web App
Flask app: upload Gronthee XLSX → review & fix OCR errors → dedup check →
process via clean_catalog.py → one-click import to Koha.
"""

import json, logging, os, re, secrets, shlex, shutil, sqlite3, subprocess, tempfile, time, uuid
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

import openpyxl
from flask import (Flask, flash, jsonify, redirect, render_template,
                   request, send_file, session, url_for)

import catalog_engine

# ── Configuration ──────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DB_PATH     = BASE_DIR / 'dedup_registry.db'
SESSIONS_DIR = BASE_DIR / 'sessions'
UPLOADS_DIR  = BASE_DIR / 'uploads'
OUTPUT_DIR   = BASE_DIR / 'output'

for d in (SESSIONS_DIR, UPLOADS_DIR, OUTPUT_DIR):
    d.mkdir(exist_ok=True)

CATALOG_PASSWORD  = os.environ.get('CATALOG_PASSWORD', 'changeme')
KOHA_INSTANCE     = os.environ.get('KOHA_INSTANCE', 'dishari_lib')
KOHA_MATCH_RULE   = os.environ.get('KOHA_MATCH_RULE', 'STRICT_CLE')
KOHA_LABEL_TEMPLATE_ID = int(os.environ.get('KOHA_LABEL_TEMPLATE_ID', '1'))   # Avery 5160 | 1 x 2-5/8
KOHA_LABEL_LAYOUT_ID   = int(os.environ.get('KOHA_LABEL_LAYOUT_ID',   '17'))  # Dishari Label
LABEL_SCRIPT        = BASE_DIR / 'create_label_batch.pl'

# Column indices — must match clean_catalog.py
COL_ISBN        = 0
COL_LANG        = 1
COL_AUTHOR      = 2
COL_TITLE       = 3
COL_SUBTITLE    = 4
COL_PAGES       = 10
COL_EDITION     = 6
COL_YEAR        = 9
COL_PUBLISHER   = 8
COL_PLACE       = 7
COL_CALL_NO     = 34
COL_ITEM_TYPE   = 25
COL_COLLECTION  = 27
COL_BRANCH_HOME = 28
COL_BRANCH_HOLD = 29
COL_DATE        = 31
COL_COST        = 33
COL_BARCODE     = 35
COL_COPY2       = 37
COL_COPY3       = 38
COL_COPY4       = 39

COPY_PREFIX = {'copy2': '11', 'copy3': '12', 'copy4': '13'}

# ── App init ───────────────────────────────────────────────────────────────
app = Flask(__name__)
_secret = os.environ.get('FLASK_SECRET_KEY', '')
if not _secret:
    _secret = secrets.token_hex(32)
    logging.warning(
        'FLASK_SECRET_KEY not set — generated a random key. '
        'All sessions will be lost on restart. '
        'Set FLASK_SECRET_KEY in your systemd unit file.'
    )
app.secret_key = _secret
app.permanent_session_lifetime = timedelta(hours=4)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024   # 10 MB hard limit (S1)

MAX_UPLOAD_ROWS = 500   # P2: reject files with more rows than this


@app.errorhandler(413)
def request_entity_too_large(_):
    flash('File too large — maximum upload size is 10 MB.')
    return redirect(url_for('index'))


# ── Database ───────────────────────────────────────────────────────────────
def init_db():
    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS books (
                id            INTEGER PRIMARY KEY,
                isbn          TEXT,
                title_norm    TEXT NOT NULL,
                author_norm   TEXT NOT NULL,
                title_display TEXT,
                author_display TEXT,
                barcode       TEXT NOT NULL,
                copies        INTEGER DEFAULT 1,
                added_at      DATETIME DEFAULT CURRENT_TIMESTAMP,
                source_file   TEXT
            )
        ''')
        # Add columns to existing DBs that predate this schema
        for col, coltype in [('title_display', 'TEXT'), ('author_display', 'TEXT'),
                              ('publisher', 'TEXT'), ('year', 'TEXT'),
                              ("edition_norm", "TEXT NOT NULL DEFAULT ''")]:
            try:
                conn.execute(f'ALTER TABLE books ADD COLUMN {col} {coltype}')
            except Exception:
                pass
        # G1: drop old 2-column index and replace with 3-column (title+author+edition)
        conn.execute('DROP INDEX IF EXISTS idx_dedup')
        conn.execute(
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_dedup '
            'ON books(title_norm, author_norm, edition_norm)'
        )
        conn.execute(
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_barcode '
            'ON books(barcode)'
        )
        # S2: persist login lockout state across restarts
        conn.execute('''
            CREATE TABLE IF NOT EXISTS login_attempts (
                ip    TEXT PRIMARY KEY,
                count INTEGER DEFAULT 1,
                since REAL NOT NULL
            )
        ''')

init_db()


def normalize(text: str) -> str:
    """Normalize text for dedup: lowercase, strip punctuation, collapse spaces."""
    if not text:
        return ''
    text = text.lower().strip()
    text = re.sub(r'[^\w\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def normalize_author(text: str) -> str:
    """
    Word-order-independent author normalization.
    Handles inverted names from clean_catalog.py:
      "Ashapurna Devi"  → "ashapurna devi"
      "Devi, Ashapurna" → "ashapurna devi"  (comma stripped, words sorted)
    Both map to the same registry key.
    """
    return ' '.join(sorted(normalize(text).split()))


_EDITION_WORD_MAP = {
    'first': '1', 'second': '2', 'third': '3', 'fourth': '4',
    'fifth': '5', 'sixth': '6', 'seventh': '7', 'eighth': '8',
    'ninth': '9', 'tenth': '10',
}
_EDITION_NOISE = re.compile(r'\b(edition|editions|ed|eds)\b\.?', re.IGNORECASE)


def normalize_edition(text: str) -> str:
    """
    Smart edition normalization so common variants map to the same key.

    Numeric ordinals (any number):
      "1", "1st", "1st Ed", "1st Edition."  → "1"
      "2nd", "2nd Ed", "2nd Edition"        → "2"
      "3rd", "3rd Ed.", "3rd Edition"       → "3"
      "4th Edition"                          → "4"

    OCR artefact — scanner misreads '1' as '!':
      "!st Ed."  (scanned copy of "1st Ed.")  → "1"

    Word ordinals:
      "First Edition", "Second Ed", "Third"  → "1", "2", "3"

    Other:
      "Revised Edition" → "revised"
      "" / None         → ""

    Steps:
      1. OCR fix: replace '!' before ordinal suffix with '1' ("!st" → "1st")
      2. Run through normalize() — lowercase, strip punctuation, collapse spaces
      3. Numeric ordinal prefix ("1st", "2nd", "3", "4th") → bare digit
      4. Word ordinal ("first", "second", …) → bare digit
      5. Strip edition/ed noise words and return remainder
    """
    if not text:
        return ''
    s = text.strip()
    # Fix OCR artefact: '!' at start of ordinal ("!st Ed" → "1st Ed")
    s = re.sub(r'!(?=st|nd|rd|th|\s|$)', '1', s, flags=re.IGNORECASE)
    s = normalize(s)   # lowercase, strip punctuation, collapse spaces
    if not s:
        return ''
    # Numeric ordinal prefix: "1st", "2nd", "3", "4th" → "1", "2", "3", "4"
    m = re.match(r'^(\d+)(?:st|nd|rd|th)?\b', s)
    if m:
        return m.group(1)
    # Word ordinal: "first edition", "second" → "1", "2"
    first_word = s.split()[0]
    if first_word in _EDITION_WORD_MAP:
        return _EDITION_WORD_MAP[first_word]
    # Strip noise words ("edition", "ed") and return whatever remains
    s = _EDITION_NOISE.sub('', s).strip()
    return s


def clean_isbn(isbn: str) -> str:
    """
    Normalize an ISBN to its ISBN-13 form, mirroring clean_catalog.py.

    Steps:
      1. Strip Excel float artifact: '8170669677.0' → '8170669677'
      2. Strip any 'ISBN' prefix text and non-digit/X characters
      3. ISBN-10 → ISBN-13: prepend '978', recalculate check digit
      4. ISBN-13: return as-is
      5. Anything else (Na, N/A, blank, unknown length): return ''

    This ensures registry lookups match regardless of whether Gronthee
    exported ISBN-10 or ISBN-13 for the same book.
    """
    s = (isbn or '').strip()
    # Remove trailing .0 / .00 from Excel float conversion
    s = re.sub(r'\.0+$', '', s)
    # Strip 'ISBN' prefix (case-insensitive) and keep only digits + X
    s = re.sub(r'(?i)isbn', '', s)
    s = re.sub(r'[^\dXx]', '', s).upper()

    if len(s) == 13 and s.isdigit():
        return s

    if len(s) == 10:
        # Convert ISBN-10 → ISBN-13 (978 prefix, new check digit)
        isbn12 = '978' + s[:9]
        try:
            total = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(isbn12))
            check = (10 - (total % 10)) % 10
            return isbn12 + str(check)
        except ValueError:
            # Non-numeric digit (e.g. X in positions 0–8) — return stripped form
            return s

    return s  # unknown length — return stripped, let lookup decide


def next_copy_action(copies: int) -> str:
    """
    Given the number of copies already in the registry, return the action to
    pre-select in the review dropdown.
      1 copy  (10xxxx only)  → add copy 2  (11xxxx)
      2 copies               → add copy 3  (12xxxx)
      3 copies               → add copy 4  (13xxxx)
      4+ copies              → skip (barcode scheme maxes out at 4 copies)
    """
    return {1: 'copy2', 2: 'copy3', 3: 'copy4'}.get(copies, 'skip')


def clean_year(year: str) -> str:
    """
    Mirror clean_catalog.py: extract first 4-digit year, validate 1800–current year.
    '1993 (reprint)' → '1993'
    '9999' → ''  (out of range)
    '1993.0' → '1993'
    """
    s = re.sub(r'[^\d]', '', (year or '').strip())[:4]
    if s.isdigit() and 1800 <= int(s) <= datetime.now().year:
        return s
    return ''


# English volume markers — run on normalize(title) where Bengali marks are stripped.
_VOL_RE = re.compile(
    r'\b(vol(?:ume)?|part|pt|khand(?:a)?|no|episode|chapter)\s*\.?\s*'
    r'(\d+|i{1,3}|iv|vi{0,3}|ix|x{1,3})\b',
    re.IGNORECASE
)
# Bengali volume markers — run on the RAW title because normalize() strips
# combining marks (virama ্, vowel signs) that are part of these words.
# Covers: খণ্ড (khanda), ভাগ (bhag), পর্ব (parba), সংখ্যা (sankhya)
_BN_VOL_RE = re.compile(
    r'(খণ্ড|ভাগ|পর্ব|সংখ্যা)\s*([০-৯\d]+)'
)
_ROMAN = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4, 'v': 5,
          'vi': 6, 'vii': 7, 'viii': 8, 'ix': 9, 'x': 10}
# Translate Bengali digits → ASCII digits for comparison
_BN_DIGITS = str.maketrans('০১২৩৪৫৬৭৮৯', '0123456789')


def extract_volume(title: str):
    """
    Return a normalised volume number string if the title contains a volume/part
    indicator, else None.
      "Sharadindu Omnibus Vol 2"  → "2"
      "Sharadindu Omnibus Vol II" → "2"  (roman → arabic)
      "রবীন্দ্র রচনাবলী খণ্ড ২"   → "2"  (Bengali, raw title)
      "Na Hanyate"                → None
    """
    # English path — search on normalized title
    m = _VOL_RE.search(normalize(title))
    if m:
        val = m.group(2).lower()
        return str(_ROMAN.get(val, val))
    # Bengali path — search on raw title (combining marks preserved)
    m = _BN_VOL_RE.search(title)
    if m:
        return m.group(2).translate(_BN_DIGITS)
    return None


def _volumes_conflict(title_a: str, title_b: str) -> bool:
    """
    Return True if the two titles have volume indicators that differ.
    If either title has NO volume indicator, no conflict is declared
    (we let the fuzzy score decide).
    """
    va, vb = extract_volume(title_a), extract_volume(title_b)
    if va is None and vb is None:
        return False        # neither has a volume marker — no conflict
    return va != vb         # one or both have markers and they differ


def lookup_dup(isbn: str, title: str, author: str, edition: str = '') -> dict | None:
    """
    Check dedup registry. Three-stage lookup:
      1. ISBN exact match (most reliable)
      2. Normalized title + author + edition exact match (G1: edition-aware)
      3. Fuzzy title + author match (catches romanization variants)
         — skipped if volumes conflict (Vol 1 ≠ Vol 2)
         — skipped if both editions are known and differ

    Returns dict with keys: barcode, title_norm, copies, fuzzy (bool).
    Returns None if no match.
    """
    isbn_clean = clean_isbn(isbn)
    ne = normalize_edition(edition)   # edition_norm for this lookup

    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row

        # 1. ISBN exact match
        if isbn_clean:
            row = conn.execute(
                'SELECT barcode, title_norm, author_norm, title_display, edition_norm, copies '
                'FROM books WHERE isbn=?',
                (isbn_clean,)
            ).fetchone()
            if row:
                return {**dict(row), 'fuzzy': False}

        # 2. Normalized title + author + edition exact match
        nt = normalize(title)
        na = normalize_author(author)
        if nt and na:
            row = conn.execute(
                'SELECT barcode, title_norm, author_norm, title_display, edition_norm, copies '
                'FROM books WHERE title_norm=? AND author_norm=? AND edition_norm=?',
                (nt, na, ne)
            ).fetchone()
            if row:
                return {**dict(row), 'fuzzy': False}

        # 3. Fuzzy fallback — only when rapidfuzz is available
        if not (nt and na):
            return None
        try:
            from rapidfuzz import fuzz as _fuzz
        except ImportError:
            return None

        TITLE_THRESHOLD  = 80   # title similarity %
        AUTHOR_THRESHOLD = 72   # author similarity % (lower: romanization varies more)

        # P1: Pre-filter candidates using the longest title word (> 3 chars) as a
        # LIKE anchor before running rapidfuzz. For an 80%+ similar title the longest
        # word almost certainly appears in both. Reduces O(N) scan to a small subset.
        sig_words = sorted([w for w in nt.split() if len(w) > 3], key=len, reverse=True)
        if sig_words:
            candidates = conn.execute(
                'SELECT barcode, title_norm, author_norm, title_display, edition_norm, copies '
                'FROM books WHERE title_norm LIKE ?',
                (f'%{sig_words[0]}%',)
            ).fetchall()
        else:
            candidates = conn.execute(
                'SELECT barcode, title_norm, author_norm, title_display, edition_norm, copies '
                'FROM books'
            ).fetchall()

        best_score, best_row = 0.0, None
        for cand in candidates:
            # Hard block: different volumes of the same series are distinct books.
            cand_title_for_vol = cand['title_display'] or cand['title_norm']
            if _volumes_conflict(title, cand_title_for_vol):
                continue

            # G1: Hard block — if both editions are known and differ, it's a different
            # edition of the same work, not a duplicate copy.
            cand_ne = cand['edition_norm'] or ''
            if ne and cand_ne and ne != cand_ne:
                continue

            t_score = _fuzz.token_sort_ratio(nt, cand['title_norm'])
            if t_score < TITLE_THRESHOLD:
                continue

            a_score = _fuzz.token_sort_ratio(na, cand['author_norm'])
            if a_score < AUTHOR_THRESHOLD:
                continue

            # Combined score weighted toward title
            combined = t_score * 0.65 + a_score * 0.35
            if combined > best_score:
                best_score = combined
                best_row   = cand

        if best_row:
            return {**dict(best_row), 'fuzzy': True, 'fuzzy_score': round(best_score)}

    return None


def register_books(books: list, source_file: str) -> tuple[int, str]:
    """
    Insert newly processed books into the dedup registry.
    Returns (skipped_count, warnings) where warnings is a newline-joined string
    of any barcode mismatches detected (same title+author but different barcode
    in registry vs. this run — usually means a prior dry-run used a different barcode).
    """
    skipped = 0
    warnings = []
    with sqlite3.connect(str(DB_PATH)) as conn:
        for b in books:
            isbn = clean_isbn(b.get('isbn') or '') or None
            nt = normalize(b['title'])
            na = normalize_author(b['author'])
            ne = normalize_edition(b.get('edition') or '')
            cur = conn.execute(
                'INSERT OR IGNORE INTO books '
                '(isbn, title_norm, author_norm, edition_norm, title_display, author_display, '
                'publisher, year, barcode, source_file) '
                'VALUES (?,?,?,?,?,?,?,?,?,?)',
                (isbn or None, nt, na, ne,
                 b['title'],
                 b['author'],
                 b.get('publisher') or None,
                 b.get('year') or None,
                 b['barcode'],
                 source_file)
            )
            if cur.rowcount == 0:
                skipped += 1
                # Check whether the stored barcode differs from what this run produced.
                stored = conn.execute(
                    'SELECT barcode FROM books WHERE title_norm=? AND author_norm=? AND edition_norm=?',
                    (nt, na, ne)
                ).fetchone()
                stored_bc = stored[0] if stored else None
                if stored_bc and stored_bc != b['barcode']:
                    msg = (
                        f'Barcode mismatch for "{b.get("title","?")}": '
                        f'registry has {stored_bc}, this run produced {b["barcode"]}. '
                        f'Registry NOT updated — verify which barcode is in Koha.'
                    )
                    warnings.append(msg)
                    logging.warning('register_books: %s', msg)
                else:
                    logging.info(
                        'register_books: skipped "%s" (barcode %s) — already in registry',
                        b.get('title', '?'), b.get('barcode', '?')
                    )
    return skipped, '\n'.join(warnings)



# ── Session helpers ────────────────────────────────────────────────────────
def session_path(sid: str) -> Path:
    return SESSIONS_DIR / f'{sid}.json'


def load_meta() -> dict:
    """Load koha_session_meta.json to get synonym dictionaries."""
    p = catalog_engine.META_FILE
    if p.exists():
        return json.loads(p.read_text(encoding='utf-8'))
    return {}


def prenormalize_author(raw: str, meta: dict) -> str:
    """
    Apply author synonym lookup + name inversion to mirror clean_catalog.py.
    "Ashapurna Devi"    → "Devi, Ashapurna"
    "sunil ganguly"     → "Gangopadhyay, Sunil"  (via synonyms_author)
    Already-inverted names ("Devi, Ashapurna") are left unchanged.

    Mirrors clean_catalog.py match_author_synonym(): exact all-words match first,
    then rapidfuzz token_sort_ratio fallback at fuzzy_author_threshold (default 88).
    """
    raw = raw.strip()
    if not raw:
        return raw

    synonyms_author = meta.get('synonyms_author', {})
    fuzzy_threshold = meta.get('fuzzy_author_threshold', 88)

    # Stage 1: exact all-words match (case-insensitive) — fast path
    raw_lower = raw.lower()
    matched = raw
    for standard, variations in synonyms_author.items():
        for v in variations:
            if all(w in raw_lower for w in v.lower().split()):
                matched = standard
                break
        else:
            continue
        break

    # Stage 2: rapidfuzz fallback — catches single-character typos exact match misses
    if matched == raw:
        try:
            from rapidfuzz import fuzz as _rf_fuzz, process as _rf_process
            candidates = {}
            for standard, variations in synonyms_author.items():
                for v in variations:
                    candidates[v.lower()] = standard
                candidates[standard.lower()] = standard
            if candidates:
                result = _rf_process.extractOne(
                    raw_lower, list(candidates.keys()),
                    scorer=_rf_fuzz.token_sort_ratio
                )
                if result and result[1] >= fuzzy_threshold:
                    matched = candidates[result[0]]
        except ImportError:
            pass

    # Inversion: skip if already comma-separated or multi-author
    if ',' in matched:
        return matched
    if re.search(r'\s+and\s+|\s*&\s*', matched, flags=re.IGNORECASE):
        return matched
    # Remove trailing "et al."
    matched = re.sub(r'\s+et\.?\s+al\.?$', '', matched, flags=re.IGNORECASE).strip()
    parts = matched.split()
    if len(parts) > 1:
        return f"{parts[-1]}, {' '.join(parts[:-1])}"
    return matched


def prenormalize_title(raw: str, meta: dict) -> str:
    """
    Apply synonyms_keywords substitutions to title/subtitle/series — mirrors clean_catalog.py.
    "Golpo Samagra" → "Galpa Samagra"
    """
    result = raw
    for standard, pattern in meta.get('synonyms_keywords', {}).items():
        result = re.sub(rf'\b({pattern})\b', standard, result, flags=re.IGNORECASE)
    return result


def prenormalize_publisher(raw: str, meta: dict) -> str:
    """
    Apply synonyms_publisher: word-boundary match against variation list → canonical name.
    Mirrors clean_catalog.py lines 352-355.
    """
    for standard, variations in meta.get('synonyms_publisher', {}).items():
        if any(re.search(rf'\b{re.escape(v)}\b', raw, flags=re.IGNORECASE) for v in variations):
            return standard
    return raw


def prenormalize_place(raw: str, meta: dict) -> str:
    """
    Apply synonyms_place: exact case-insensitive match against variations list.
    Mirrors clean_catalog.py lines 402-406.
    """
    raw_lower = raw.lower()
    for standard, variations in meta.get('synonyms_place', {}).items():
        if raw_lower in [v.lower() for v in variations]:
            return standard
    return raw


def prenormalize_subject(raw: str, meta: dict) -> str:
    """
    Apply synonyms_subject: word-boundary regex match → canonical heading.
    Mirrors clean_catalog.py lines 561-565.
    """
    for standard, variations in meta.get('synonyms_subject', {}).items():
        if any(re.search(rf'\b{re.escape(v)}\b', raw, flags=re.IGNORECASE) for v in variations):
            return standard
    return raw


def prenormalize_series(raw: str, author_pre_invert: str, meta: dict) -> str:
    """
    Apply synonyms_keywords to series text, then apply series_overrides keyed by
    "SeriesTitle|AuthorName" (author before inversion, matching clean_catalog.py line 446).
    """
    normalized = prenormalize_title(raw, meta)
    key = f"{normalized}|{author_pre_invert}"
    return meta.get('series_overrides', {}).get(key, normalized)


def _first_author_for_dedup(raw: str, meta: dict) -> str:
    """
    Extract the first author from a potentially multi-author string and prenormalize it.
    Mirrors clean_catalog.py: "Sunil Ganguly and Tapas Biswas" → prenormalize "Sunil Ganguly".

    The full raw string is kept in cols[] for MARC generation; this function is only
    used to produce the lookup key for dedup registry queries.
    """
    # Split on "and" / "&" the same way clean_catalog.py does
    first = re.split(r'\s+and\s+|\s*&\s*', raw, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    return prenormalize_author(first, meta)


def load_session(sid: str) -> dict | None:
    p = session_path(sid)
    if not p.exists():
        return None
    return json.loads(p.read_text(encoding='utf-8'))


def save_session(sid: str, data: dict):
    session_path(sid).write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8'
    )


# ── Auth decorator ────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('auth'):
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return wrapper


# ── File parsing ───────────────────────────────────────────────────────────
def parse_upload(filepath: str) -> list[list]:
    """
    Parse an uploaded XLSX or CSV into a list of rows (each row = list of 40+ values).
    Returns rows as lists of strings; pads to at least 40 columns.
    """
    path = Path(filepath)
    if path.suffix.lower() in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header
            vals = [str(v).strip() if v is not None else '' for v in row]
            # Pad to 40 columns
            while len(vals) < 40:
                vals.append('')
            rows.append(vals)
        wb.close()
    else:
        import csv
        rows = []
        with open(filepath, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i == 0:
                    continue
                vals = [v.strip() for v in row]
                while len(vals) < 40:
                    vals.append('')
                rows.append(vals)
    return [r for r in rows if any(v for v in r)]  # drop fully blank rows


def build_review_rows(raw_rows: list[list], source_file: str) -> list[dict]:
    """
    Convert raw XLSX rows to review-ready dicts with initial dedup status.
    Author and title are pre-normalized (synonym + inversion) so the review
    screen shows canonical values matching what clean_catalog.py will produce.
    """
    meta   = load_meta()
    # Added-author column indices (cols 19–24) — must match clean_catalog.py COL_ADDED_AUTHORS
    COL_ADDED_AUTHORS = [19, 20, 21, 22, 23, 24]
    # Subject column indices (cols 14–18)
    COL_SUBJECTS_LIST = [14, 15, 16, 17, 18]
    COL_SUBTITLE_IDX  = 4
    COL_SERIES_IDX    = 12

    review = []
    # G3: track books seen within this upload to flag within-file duplicates
    # Keys: ('isbn', <isbn>) or ('tae', (title_norm, author_norm, edition_norm))
    # Values: 0-based row index of first occurrence
    seen_in_upload: dict = {}

    for idx, cols in enumerate(raw_rows):
        cols = list(cols)   # make mutable copy before normalizing

        # Clean raw cell values before any normalization
        cols[COL_ISBN] = clean_isbn(cols[COL_ISBN])
        cols[COL_YEAR] = clean_year(cols[COL_YEAR])
        # Strip Excel float artifact from page count: "300.0" → "300"
        if len(cols) > COL_PAGES:
            cols[COL_PAGES] = re.sub(r'\.0+$', '', str(cols[COL_PAGES] or '').strip())

        isbn = cols[COL_ISBN]

        # Author: synonym match + inversion (save raw pre-invert form for series key)
        raw_author_pre_invert = cols[COL_AUTHOR]
        author    = prenormalize_author(raw_author_pre_invert, meta)
        title     = prenormalize_title(cols[COL_TITLE], meta)
        publisher = prenormalize_publisher(cols[COL_PUBLISHER], meta)
        year      = cols[COL_YEAR]

        # Write normalized primary fields back into cols
        cols[COL_AUTHOR]    = author
        cols[COL_TITLE]     = title
        cols[COL_PUBLISHER] = publisher

        # Subtitle (col 4) — synonyms_keywords
        cols[COL_SUBTITLE_IDX] = prenormalize_title(cols[COL_SUBTITLE_IDX], meta)

        # Place (col 7) — synonyms_place
        cols[COL_PLACE] = prenormalize_place(cols[COL_PLACE], meta)

        # Series (col 12) — synonyms_keywords + series_overrides
        cols[COL_SERIES_IDX] = prenormalize_series(
            cols[COL_SERIES_IDX], raw_author_pre_invert, meta
        )

        # Subjects (cols 14–18) — synonyms_subject
        for ci in COL_SUBJECTS_LIST:
            if ci < len(cols) and cols[ci]:
                cols[ci] = prenormalize_subject(cols[ci], meta)

        # Added authors (cols 19–24) — same synonym + inversion as primary author
        for ci in COL_ADDED_AUTHORS:
            if ci < len(cols) and cols[ci]:
                cols[ci] = prenormalize_author(cols[ci], meta)

        # For dedup lookup: use first author only (mirrors clean_catalog.py which stores
        # only the primary author in the registry). The full author string in cols is
        # preserved so clean_catalog.py can still split "A and B" → 100$a + 700$a.
        author_for_dedup = _first_author_for_dedup(raw_author_pre_invert, meta)
        edition = cols[COL_EDITION] if len(cols) > COL_EDITION else ''
        dup = lookup_dup(isbn, title, author_for_dedup, edition)

        # G3: within-upload duplicate check (only when no registry match found)
        same_file_dup_row = None   # 1-indexed row number of first occurrence in this upload
        if not dup and title:
            tn = normalize(title)
            an = normalize_author(author_for_dedup)
            ne = normalize_edition(edition)
            # Check ISBN first (most reliable), then title+author+edition
            if isbn and ('isbn', isbn) in seen_in_upload:
                same_file_dup_row = seen_in_upload[('isbn', isbn)] + 1
            elif tn and an and ('tae', (tn, an, ne)) in seen_in_upload:
                same_file_dup_row = seen_in_upload[('tae', (tn, an, ne))] + 1

        fuzzy_score = None   # initialise so every branch can safely reference it
        dup_source  = None
        if not title and not author:
            status      = 'ERROR'
            action      = 'skip'
            dup_barcode = dup_title = None
        elif dup and not dup['fuzzy']:
            status      = 'DUPLICATE'
            dup_barcode = dup['barcode']
            dup_title   = dup.get('title_display') or dup['title_norm']
            action      = next_copy_action(dup.get('copies', 1))
            dup_source  = 'registry'
        elif dup and dup['fuzzy']:
            status      = 'FUZZY'
            dup_barcode = dup['barcode']
            dup_title   = dup.get('title_display') or dup['title_norm']
            fuzzy_score = dup.get('fuzzy_score')
            action      = 'review'  # force reviewer to make an explicit choice
            dup_source  = 'registry'
        elif same_file_dup_row is not None:
            status      = 'DUPLICATE'
            dup_barcode = None          # no barcode yet — not in registry
            dup_title   = title         # same title by definition
            action      = 'skip'        # default; reviewer can change to copy2/3/4
            dup_source  = 'upload'
        else:
            status      = 'NEW'
            action      = 'skip'
            dup_barcode = dup_title = fuzzy_score = None

        # Register this row's dedup keys so later rows in the same upload can match it.
        # We always register (even DUPLICATE rows) so the first seen copy is the anchor.
        if title:
            tn = normalize(title)
            an = normalize_author(author_for_dedup)
            ne = normalize_edition(edition)
            if isbn:
                seen_in_upload.setdefault(('isbn', isbn), idx)
            if tn and an:
                seen_in_upload.setdefault(('tae', (tn, an, ne)), idx)

        review.append({
            'idx':          idx,
            'status':       status,
            'action':       action,
            'dup_barcode':  dup_barcode,
            'dup_title':    dup_title,
            'dup_source':   dup_source,
            'dup_row_num':  same_file_dup_row,
            'fuzzy_score':  fuzzy_score,
            'isbn':         isbn,
            'title':        title,
            'author':       author,
            'year':         year,
            'publisher':    publisher,
            'subtitle':     cols[COL_SUBTITLE]   if len(cols) > COL_SUBTITLE   else '',
            'pages':        cols[COL_PAGES]      if len(cols) > COL_PAGES      else '',
            'category':     cols[14]             if len(cols) > 14             else '',
            'genre':        cols[15]             if len(cols) > 15             else '',
            'item_type':    cols[COL_ITEM_TYPE]  if len(cols) > COL_ITEM_TYPE  else '',
            'ccode':        cols[COL_COLLECTION] if len(cols) > COL_COLLECTION else '',
            'cols':         cols,
        })
    return review


# ── Koha item addition ─────────────────────────────────────────────────────
def add_copy_item_to_koha(dup_barcode: str, copy_barcode: str, copy_num: int, meta: dict) -> tuple[bool, str]:
    """
    Add a copy item to an existing Koha bib via direct MySQL insertion.

    Reads Koha MySQL credentials from koha-conf.xml, looks up the biblionumber
    from the existing primary barcode, then inserts a new item row.

    Args:
        dup_barcode: primary barcode of existing book (10XXXX)
        copy_barcode: new copy barcode (11/12/13XXXX)
        copy_num: copy number (2, 3, or 4)
        meta: dict with home_branch, hold_branch, item_type, call_no, date

    Returns: (success: bool, message: str)
    """
    try:
        import xml.etree.ElementTree as ET
        import mysql.connector

        # Read MySQL credentials from koha-conf.xml
        koha_conf = '/etc/koha/sites/dishari_lib/koha-conf.xml'
        tree = ET.parse(koha_conf)
        root = tree.getroot()

        def get_elem(tag):
            el = root.find(f'.//{tag}')
            return el.text.strip() if el is not None and el.text else ''

        db_host = get_elem('hostname') or 'localhost'
        db_name = get_elem('database')
        db_user = get_elem('user')
        db_pass = get_elem('pass')

        if not (db_name and db_user):
            return False, 'Could not read Koha MySQL credentials from koha-conf.xml'

        # Connect to MySQL
        cnx = mysql.connector.connect(
            host=db_host,
            user=db_user,
            password=db_pass,
            database=db_name
        )
        cursor = cnx.cursor()

        # Look up existing item by primary barcode
        cursor.execute('SELECT biblionumber, biblioitemnumber FROM items WHERE barcode = %s', (dup_barcode,))
        row = cursor.fetchone()
        if not row:
            return False, f'No item found with barcode {dup_barcode}'

        biblionumber, biblioitemnumber = row

        # Insert new item
        home_branch = meta.get('home_branch', 'DFL')
        hold_branch = meta.get('hold_branch', 'DFL')
        item_type = meta.get('item_type', 'BK')
        call_no = meta.get('call_no', '891')
        date_str = meta.get('date', '')

        cursor.execute('''
            INSERT INTO items
            (biblionumber, biblioitemnumber, barcode, homebranch, holdingbranch,
             itype, itemcallnumber, dateaccessioned, copynumber)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (biblionumber, biblioitemnumber, copy_barcode, home_branch, hold_branch,
              item_type, call_no, date_str, copy_num))

        cnx.commit()
        cursor.close()
        cnx.close()

        return True, f'Copy item added: barcode {copy_barcode}'

    except Exception as exc:
        return False, f'Error adding copy item: {str(exc)}'


# ── Koha import ────────────────────────────────────────────────────────────
def import_to_koha(mrc_path: str) -> tuple[bool, str]:
    """
    Import MARC file via bulkmarcimport.pl on the Koha server.

    Copies the file to /tmp first (Koha user can't access /home/dishari).
    Uses --insert to add new bibs only. COPY records are handled separately
    via add_copy_item_to_koha() which directly inserts items via Koha API.

    Returns (success, message).
    """
    import shutil
    if not shutil.which('koha-shell'):
        return False, 'koha-shell not found — not running on the Koha server.'
    try:
        # Copy to /tmp so Koha user can read it (home dir is not accessible)
        tmp_mrc = Path('/tmp') / f'catalog_import_{Path(mrc_path).stem}.mrc'
        shutil.copy2(mrc_path, tmp_mrc)

        try:
            cmd = [
                'sudo', 'koha-shell', KOHA_INSTANCE, '-c',
                f'/usr/share/koha/bin/migration_tools/bulkmarcimport.pl -b --insert -file {shlex.quote(str(tmp_mrc))}'
            ]
            result = subprocess.run(cmd, capture_output=True, text=True,
                                   stdin=subprocess.DEVNULL, timeout=180)
            if result.returncode == 0:
                return True, result.stdout or 'Import completed.'
            return False, f'Import exited {result.returncode}:\n{result.stderr}'
        finally:
            # Clean up temp file
            tmp_mrc.unlink(missing_ok=True)
    except FileNotFoundError:
        return False, 'koha-shell not found — not running on the Koha server.'
    except subprocess.TimeoutExpired:
        return False, 'Import timed out after 3 minutes.'
    except Exception as exc:
        return False, str(exc)


def rollback_registry(source_file: str) -> int:
    """
    Delete all registry entries from a failed import session.
    Returns count of rows deleted.
    """
    with sqlite3.connect(str(DB_PATH)) as conn:
        cur = conn.execute('DELETE FROM books WHERE source_file=?', (source_file,))
        deleted = cur.rowcount
        conn.commit()
        if deleted > 0:
            logging.info(f'Rolled back {deleted} registry entries for source_file={source_file}')
    return deleted


def create_label_pdf(barcodes: list[str], pdf_path: str) -> tuple[bool, str]:
    """
    Create a Koha label batch and export it as a PDF via koha-shell.
    The PDF is written to pdf_path by the Perl script.
    Returns (success, message).
    Only callable on the Koha server where koha-shell is present.
    """
    if not shutil.which('koha-shell'):
        return False, 'koha-shell not found — not running on the Koha server.'
    if not LABEL_SCRIPT.exists():
        return False, f'{LABEL_SCRIPT.name} not found in app directory.'

    barcode_str = ','.join(b for b in barcodes if b)
    if not barcode_str:
        return False, 'No barcodes provided.'

    try:
        cmd = [
            'sudo', 'koha-shell', KOHA_INSTANCE, '-c',
            f'perl {shlex.quote(str(LABEL_SCRIPT))} '
            f'{shlex.quote(barcode_str)} '
            f'{KOHA_LABEL_TEMPLATE_ID} '
            f'{KOHA_LABEL_LAYOUT_ID} '
            f'{shlex.quote(pdf_path)}'
        ]
        res = subprocess.run(cmd, capture_output=True, text=True,
                             stdin=subprocess.DEVNULL, timeout=60)
        if res.returncode != 0:
            return False, (res.stderr or res.stdout or 'Script exited non-zero').strip()
        # Result line (batch_id:items_added) is written to STDERR by the Perl script
        # because STDOUT is redirected to the PDF file.
        return True, res.stderr.strip()
    except subprocess.TimeoutExpired:
        return False, 'Label PDF generation timed out.'
    except Exception as exc:
        return False, str(exc)


# ── CSRF protection ────────────────────────────────────────────────────────
def _get_csrf_token() -> str:
    """Return (creating if needed) the per-session CSRF token."""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(16)
    return session['csrf_token']

def _validate_csrf():
    """Abort 403 if the submitted CSRF token doesn't match the session."""
    token = (request.form.get('csrf_token') or
             request.headers.get('X-CSRF-Token', ''))
    if not token or token != session.get('csrf_token'):
        from flask import abort
        abort(403, 'Invalid or missing CSRF token.')

# Make csrf_token() callable from every Jinja template
app.jinja_env.globals['csrf_token'] = _get_csrf_token


# ── Login rate limiting (S2: DB-backed — survives restarts) ───────────────
_MAX_FAILURES    = 5
_LOCKOUT_SECONDS = 900        # 15 minutes

def _is_rate_limited(ip: str) -> bool:
    with sqlite3.connect(str(DB_PATH)) as conn:
        row = conn.execute(
            'SELECT count, since FROM login_attempts WHERE ip=?', (ip,)
        ).fetchone()
    if not row:
        return False
    count, since = row
    if time.time() - since > _LOCKOUT_SECONDS:
        with sqlite3.connect(str(DB_PATH)) as conn:
            conn.execute('DELETE FROM login_attempts WHERE ip=?', (ip,))
        return False
    return count >= _MAX_FAILURES

def _record_login_failure(ip: str):
    now = time.time()
    with sqlite3.connect(str(DB_PATH)) as conn:
        row = conn.execute(
            'SELECT count, since FROM login_attempts WHERE ip=?', (ip,)
        ).fetchone()
        if row and now - row[1] <= _LOCKOUT_SECONDS:
            conn.execute(
                'UPDATE login_attempts SET count=count+1 WHERE ip=?', (ip,)
            )
        else:
            conn.execute(
                'INSERT OR REPLACE INTO login_attempts (ip, count, since) VALUES (?,?,?)',
                (ip, 1, now)
            )

def _clear_login_failures(ip: str):
    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.execute('DELETE FROM login_attempts WHERE ip=?', (ip,))


# ── Disk cleanup ───────────────────────────────────────────────────────────
_last_cleanup: float = 0.0

def _cleanup_old_files():
    """Delete uploads/outputs/sessions older than 7 days, and stale temp dirs."""
    global _last_cleanup
    now = time.time()
    if now - _last_cleanup < 86400:   # run at most once per 24 h
        return
    _last_cleanup = now
    cutoff      = now - 7 * 86400    # 7-day retention for user data
    cutoff_tmp  = now - 86400         # 1-day retention for temp dirs
    for directory in (UPLOADS_DIR, OUTPUT_DIR, SESSIONS_DIR):
        for p in directory.iterdir():
            try:
                if p.stat().st_mtime < cutoff:
                    p.unlink(missing_ok=True)
            except Exception:
                pass
    tmp = Path(tempfile.gettempdir())
    for p in tmp.glob('catalog_run_*'):
        try:
            if p.stat().st_mtime < cutoff_tmp:
                shutil.rmtree(str(p), ignore_errors=True)
        except Exception:
            pass


# ── Routes ─────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    ip = request.remote_addr or '0.0.0.0'
    if request.method == 'POST':
        if _is_rate_limited(ip):
            flash('Too many failed attempts — please wait 15 minutes.')
            return render_template('login.html'), 429
        _validate_csrf()
        if request.form.get('password') == CATALOG_PASSWORD:
            _clear_login_failures(ip)
            session['auth'] = True
            session.permanent = True
            return redirect(request.args.get('next') or url_for('index'))
        _record_login_failure(ip)
        flash('Incorrect password.')
    return render_template('login.html')


@app.route('/health')
def health():
    """Uptime / readiness probe — no auth required."""
    try:
        with sqlite3.connect(str(DB_PATH)) as conn:
            conn.execute('SELECT 1 FROM books LIMIT 1')
        db_ok = True
    except Exception:
        db_ok = False
    status = 200 if db_ok else 503
    return jsonify(status='ok' if db_ok else 'degraded', db=db_ok), status


@app.route('/heartbeat')
@login_required
def heartbeat():
    """Touch the session to keep auth cookie alive; called by review page JS."""
    return jsonify(ok=True)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return render_template('upload.html')


@app.route('/upload', methods=['POST'])
@login_required
def upload():
    _validate_csrf()
    _cleanup_old_files()   # opportunistic daily cleanup
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Please select a file.')
        return redirect(url_for('index'))

    ext = Path(f.filename).suffix.lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        flash('Only .xlsx / .xls / .csv files are accepted.')
        return redirect(url_for('index'))

    sid = str(uuid.uuid4())[:8]
    save_path = str(UPLOADS_DIR / f'{sid}{ext}')
    f.save(save_path)

    try:
        raw_rows = parse_upload(save_path)
    except Exception as exc:
        flash(f'Could not parse file: {exc}')
        return redirect(url_for('index'))

    if not raw_rows:
        flash('File appears to be empty or has no data rows.')
        return redirect(url_for('index'))

    if len(raw_rows) > MAX_UPLOAD_ROWS:
        flash(
            f'File has {len(raw_rows)} rows — maximum allowed is {MAX_UPLOAD_ROWS}. '
            f'Split the file into smaller batches and upload separately.'
        )
        Path(save_path).unlink(missing_ok=True)
        return redirect(url_for('index'))

    rows = build_review_rows(raw_rows, f.filename)
    save_session(sid, {
        'filename':    f.filename,
        'upload_path': save_path,
        'uploaded_at': datetime.now().isoformat(),
        'rows':        rows,
        'result':      None,
    })

    return redirect(url_for('review', sid=sid))


@app.route('/review/<sid>')
@login_required
def review(sid):
    data = load_session(sid)
    if not data:
        flash('Session not found — please re-upload.')
        return redirect(url_for('index'))
    return render_template('review.html', sid=sid, rows=data['rows'],
                           filename=data['filename'])


@app.route('/api/dedup')
@login_required
def api_dedup():
    """AJAX endpoint: check a single title/author/isbn against the registry.

    Applies the same pre-normalization as build_review_rows() so that synonym
    substitutions and author inversion are consistent between initial load and
    live AJAX re-checks triggered by user edits.
    """
    isbn    = clean_isbn(request.args.get('isbn', ''))
    title   = request.args.get('title', '')
    author  = request.args.get('author', '')
    edition = request.args.get('edition', '')
    meta    = load_meta()
    # Mirror build_review_rows(): normalize title and author before lookup
    title  = prenormalize_title(title, meta)
    # For multi-author strings, extract first author for dedup (same as clean_catalog.py)
    author = _first_author_for_dedup(author, meta)
    dup = lookup_dup(isbn, title, author, edition)
    if dup and not dup['fuzzy']:
        return jsonify(status='DUPLICATE',
                       dup_barcode=dup['barcode'],
                       dup_title=dup.get('title_display') or dup['title_norm'],
                       next_action=next_copy_action(dup.get('copies', 1)))
    if dup and dup['fuzzy']:
        return jsonify(status='FUZZY',
                       dup_barcode=dup['barcode'],
                       dup_title=dup.get('title_display') or dup['title_norm'],
                       fuzzy_score=dup.get('fuzzy_score'),
                       next_action='review')
    if not title.strip() and not author.strip():
        return jsonify(status='ERROR', dup_barcode=None, dup_title=None, next_action=None)
    return jsonify(status='NEW', dup_barcode=None, dup_title=None, next_action=None)


@app.route('/api/book')
@login_required
def api_book():
    """Return full registry details for a book identified by its primary barcode."""
    barcode = request.args.get('barcode', '').strip()
    if not barcode:
        return jsonify(error='No barcode'), 400
    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            'SELECT isbn, title_display, title_norm, author_display, author_norm, '
            'publisher, year, barcode, copies, added_at, source_file FROM books WHERE barcode=?',
            (barcode,)
        ).fetchone()
    if not row:
        return jsonify(error='Not found'), 404
    r = dict(row)
    title  = r.get('title_display')  or r.get('title_norm',  '').title()
    author = r.get('author_display') or r.get('author_norm', '').title()
    # Build list of all known barcodes for this book
    prefix = barcode[:2]            # e.g. "10"
    suffix = barcode[2:]            # e.g. "0045"
    copies_list = []
    copy_labels = {'10': 'Copy 1', '11': 'Copy 2', '12': 'Copy 3', '13': 'Copy 4'}
    for p, label in copy_labels.items():
        bc = p + suffix
        # primary barcode is stored; copies are derived
        copies_list.append({'label': label, 'barcode': bc,
                             'is_primary': p == '10'})
    copies_list = copies_list[:r.get('copies', 1)]
    return jsonify(
        title=title, author=author,
        isbn=r.get('isbn') or '',
        publisher=r.get('publisher') or '',
        year=r.get('year') or '',
        copies=r.get('copies', 1),
        copies_list=copies_list,
        added_at=(r.get('added_at') or '')[:10],
        source_file=r.get('source_file') or '',
    )


@app.route('/process/<sid>', methods=['POST'])
@login_required
def process(sid):
    _validate_csrf()
    data = load_session(sid)
    if not data:
        flash('Session expired — please re-upload.')
        return redirect(url_for('index'))

    row_count = int(request.form.get('row_count', 0))

    # Rebuild rows from form (user may have edited title/author/isbn)
    new_rows  = []
    copy_rows = []

    for i in range(row_count):
        status    = request.form.get(f'status_{i}', 'NEW')
        action    = request.form.get(f'action_{i}', 'skip')
        title     = request.form.get(f'title_{i}',     '').strip()
        author    = request.form.get(f'author_{i}',    '').strip()
        isbn      = request.form.get(f'isbn_{i}',      '').strip()
        year      = request.form.get(f'year_{i}',      '').strip()
        publisher = request.form.get(f'publisher_{i}', '').strip()
        subtitle  = request.form.get(f'subtitle_{i}',  '').strip()
        pages     = request.form.get(f'pages_{i}',     '').strip()
        category  = request.form.get(f'category_{i}',  '').strip()
        genre     = request.form.get(f'genre_{i}',     '').strip()
        item_type = request.form.get(f'item_type_{i}', '').strip()
        ccode     = request.form.get(f'ccode_{i}',     '').strip()
        edition   = request.form.get(f'col_{i}_6',    '').strip()
        dup_bc    = request.form.get(f'dup_barcode_{i}', '')

        # Rebuild the full 40-column list with user edits applied
        cols = list(data['rows'][i]['cols']) if i < len(data['rows']) else [''] * 40
        cols[COL_ISBN]      = isbn
        cols[COL_AUTHOR]    = author
        cols[COL_TITLE]     = title
        if year:      cols[COL_YEAR]      = year
        if publisher: cols[COL_PUBLISHER] = publisher
        if edition is not None: cols[COL_EDITION] = edition
        if subtitle:  cols[COL_SUBTITLE]  = subtitle
        if pages:     cols[COL_PAGES]     = pages
        if category:  cols[14]            = category
        if genre:     cols[15]            = genre
        if item_type: cols[COL_ITEM_TYPE]   = item_type
        if ccode:     cols[COL_COLLECTION]  = ccode
        # Clear copy trigger columns for all rows (will be set only for copy rows)
        cols[COL_COPY2] = ''
        cols[COL_COPY3] = ''
        cols[COL_COPY4] = ''

        row_meta = {
            'idx':         i,
            'title':       title,
            'author':      author,
            'isbn':        isbn,
            'edition':     cols[COL_EDITION],
            'year':        year      or cols[COL_YEAR],
            'publisher':   publisher or cols[COL_PUBLISHER],
            'place':       cols[COL_PLACE],
            'call_no':     cols[COL_CALL_NO],
            'item_type':   cols[COL_ITEM_TYPE],
            'ccode':       cols[COL_COLLECTION],
            'cost':        cols[COL_COST],
            'home_branch': cols[COL_BRANCH_HOME],
            'hold_branch': cols[COL_BRANCH_HOLD],
            'date':        cols[COL_DATE],
            'dup_barcode': dup_bc,
        }

        if request.form.get(f'deleted_{i}') == '1':
            continue  # user excluded this row

        if status == 'ERROR':
            continue  # blocked by front-end, but skip defensively

        if status in ('NEW', 'FUZZY_NEW'):
            # FUZZY_NEW = user decided the fuzzy match is wrong; treat as a new book
            new_rows.append({'cols': cols, 'meta': row_meta})

        elif status in ('DUPLICATE', 'FUZZY'):
            if action in ('copy2', 'copy3', 'copy4'):
                # Derive copy barcode from stored primary barcode
                if dup_bc and len(dup_bc) >= 4:
                    prefix = COPY_PREFIX[action]
                    copy_bc = prefix + dup_bc[2:]
                    copy_num = {'copy2': 2, 'copy3': 3, 'copy4': 4}[action]
                    copy_rows.append({'cols': cols, 'meta': row_meta,
                                      'action': action, 'copy_bc': copy_bc,
                                      'copy_num': copy_num})
            # else: action == 'skip', do nothing

    if not new_rows and not copy_rows:
        flash('Nothing to process — all rows were skipped or had errors.')
        return redirect(url_for('review', sid=sid))

    # ── Process new books ─────────────────────────────────────────────────
    mrc_bytes      = b''
    audit_path     = None
    errors_path    = None
    new_book_count = 0
    skipped_reg    = 0
    engine_errors  = ''
    label_barcodes = []     # accumulates all barcodes inserted this session

    if new_rows:
        tmp_xlsx = tempfile.NamedTemporaryFile(
            suffix='.xlsx', delete=False, dir=str(UPLOADS_DIR), prefix=f'{sid}_new_'
        )
        tmp_xlsx.close()
        _write_xlsx(new_rows, tmp_xlsx.name)

        eng_result = catalog_engine.run(tmp_xlsx.name)
        engine_errors = eng_result.get('stderr', '')

        if eng_result['success'] and eng_result['mrc']:
            with open(eng_result['mrc'], 'rb') as f:
                mrc_bytes = f.read()
            audit_path  = eng_result['audit']
            errors_path = eng_result['errors']

            # Register processed books in dedup registry
            if audit_path:
                processed = catalog_engine.extract_processed_books(audit_path)
                skipped_reg, reg_warnings = register_books(processed, data['filename'])
                new_book_count = len(processed)
                label_barcodes.extend(b['barcode'] for b in processed if b.get('barcode'))
                if reg_warnings:
                    engine_errors = (engine_errors + '\n' + reg_warnings).strip()

            # Copy audit/error files to persistent output dir, then remove temp dir
            if audit_path:
                dest = str(OUTPUT_DIR / f'{sid}_audit.xlsx')
                shutil.copy2(audit_path, dest)
                audit_path = dest
            if errors_path:
                dest = str(OUTPUT_DIR / f'{sid}_errors.xlsx')
                shutil.copy2(errors_path, dest)
                errors_path = dest
            if eng_result.get('out_dir'):
                shutil.rmtree(eng_result['out_dir'], ignore_errors=True)

    # ── Process copies (atomic: FileLock + SQLite EXCLUSIVE tx) ──────────────
    # Re-derives the copy number from the current DB state inside the lock so
    # that two concurrent /process requests can never assign the same barcode.
    try:
        from filelock import FileLock as _FileLock
    except ImportError:
        class _FileLock:                           # no-op fallback
            def __init__(self, p, timeout=60): pass
            def __enter__(self): return self
            def __exit__(self, *a): pass

    copy_count = 0
    for cr in copy_rows:
        try:
            meta = cr['meta']
            isbn_clean = clean_isbn(meta.get('isbn') or '')
            nt = normalize(meta['title'])
            na = normalize_author(meta['author'])
            dup_bc = meta.get('dup_barcode', '') or ''
            if not dup_bc or len(dup_bc) < 4:
                engine_errors += f'\nRow {meta["idx"]}: no primary barcode — skipped copy'
                continue

            with _FileLock(catalog_engine.LOCK_FILE, timeout=60):
                with sqlite3.connect(str(DB_PATH), timeout=10) as conn:
                    conn.execute('BEGIN EXCLUSIVE')
                    if isbn_clean:
                        row = conn.execute(
                            'SELECT copies FROM books WHERE isbn=?', (isbn_clean,)
                        ).fetchone()
                    else:
                        row = conn.execute(
                            'SELECT copies FROM books WHERE title_norm=? AND author_norm=?',
                            (nt, na)
                        ).fetchone()

                    current = row[0] if row else 1
                    next_num = current + 1
                    _prefix_map = {2: '11', 3: '12', 4: '13'}
                    if next_num not in _prefix_map:
                        engine_errors += (
                            f'\nRow {meta["idx"]}: already has {current} copies (max 4) — skipped'
                        )
                        continue

                    actual_copy_bc = _prefix_map[next_num] + dup_bc[2:]

                    # Add item to existing Koha bib via Perl API (not via MARC)
                    copy_ok, copy_msg = add_copy_item_to_koha(dup_bc, actual_copy_bc, next_num, meta)
                    if copy_ok:
                        label_barcodes.append(actual_copy_bc)
                    else:
                        engine_errors += f'\nRow {meta["idx"]}: {copy_msg}'

                    if isbn_clean:
                        conn.execute('UPDATE books SET copies=copies+1 WHERE isbn=?', (isbn_clean,))
                    else:
                        conn.execute(
                            'UPDATE books SET copies=copies+1 '
                            'WHERE title_norm=? AND author_norm=?', (nt, na)
                        )
            copy_count += 1
        except Exception as exc:
            engine_errors += f'\nCopy error row {cr["meta"]["idx"]}: {exc}'

    # ── Save merged MARC ──────────────────────────────────────────────────
    mrc_out_path = None
    if mrc_bytes:
        mrc_out_path = str(OUTPUT_DIR / f'{sid}.mrc')
        with open(mrc_out_path, 'wb') as f:
            f.write(mrc_bytes)

    # ── Attempt Koha import ───────────────────────────────────────────────
    import_ok  = False
    import_msg = ''
    if mrc_out_path:
        import_ok, import_msg = import_to_koha(mrc_out_path)
        # Auto-rollback registry if import failed
        if not import_ok and new_book_count > 0:
            rolled_back = rollback_registry(data['filename'])
            import_msg += f'\n[Rolled back {rolled_back} registry entries]'
            new_book_count = 0
            label_barcodes = []

    # ── Generate Koha label PDF (only when import succeeded) ─────────────
    labels_path = None
    if import_ok and label_barcodes:
        pdf_out = str(OUTPUT_DIR / f'{sid}_labels.pdf')
        ok, lmsg = create_label_pdf(label_barcodes, pdf_out)
        if ok and Path(pdf_out).exists():
            labels_path = pdf_out
        elif not ok:
            engine_errors = (engine_errors + f'\nLabel PDF: {lmsg}').strip()

    # ── Save result to session ────────────────────────────────────────────
    data['result'] = {
        'new_books':     new_book_count,
        'already_in_registry': skipped_reg,
        'copies_added':  copy_count,
        'skipped':       sum(1 for i in range(row_count)
                             if not request.form.get(f'deleted_{i}')
                             and request.form.get(f'status_{i}') in ('DUPLICATE', 'FUZZY')
                             and request.form.get(f'action_{i}', 'skip') == 'skip'),
        'errors':        sum(1 for i in range(row_count)
                             if not request.form.get(f'deleted_{i}')
                             and request.form.get(f'status_{i}') == 'ERROR'),
        'mrc_path':      mrc_out_path,
        'audit_path':    audit_path,
        'errors_path':   errors_path,
        'import_ok':     import_ok,
        'import_msg':    import_msg,
        'engine_errors': engine_errors,
        'labels_path':   labels_path,
    }
    save_session(sid, data)

    return redirect(url_for('result', sid=sid))


@app.route('/result/<sid>')
@login_required
def result(sid):
    data = load_session(sid)
    if not data or not data.get('result'):
        flash('No result found.')
        return redirect(url_for('index'))
    return render_template('result.html', sid=sid,
                           filename=data['filename'], r=data['result'])


@app.route('/download/<sid>/<filetype>')
@login_required
def download(sid, filetype):
    data = load_session(sid)
    if not data or not data.get('result'):
        return 'Session not found', 404

    r = data['result']
    paths = {
        'mrc':    r.get('mrc_path'),
        'audit':  r.get('audit_path'),
        'errors': r.get('errors_path'),
        'labels': r.get('labels_path'),
    }
    path = paths.get(filetype)
    if not path or not os.path.exists(path):
        return 'File not found', 404

    return send_file(path, as_attachment=True,
                     download_name=Path(path).name)


# ── Helpers ────────────────────────────────────────────────────────────────

def _write_xlsx(rows: list, out_path: str):
    """Write a list of {'cols': [...40 values...]} to an XLSX file clean_catalog.py can read."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write a minimal header row matching Gronthee column names
    headers = [
        'ISBN', 'Language', 'Author', 'Title', 'Sub Title', 'Other Title',
        'Edition', 'Publication Place', 'Publisher', 'Published Year',
        'Page Count', 'other physical details', 'Series', 'Note Area',
        'Category', 'Genre', 'Subject 3', 'Subject 4', 'Subject 5',
        'Second Author', 'Third Column', 'Editor', 'Compiler',
        'Translator', 'Illustrator', 'Item Type', 'Status', 'Collection',
        'Home Branch', 'Holding Branch', 'Shelving Location', 'Scan Date',
        'Source of Aquisition', 'Cost, normal purchase price',
        'Call No', 'BarCode', 'Public Note', 'Second Copy',
        'Third Copy', 'Fourth Copy',
    ]
    ws.append(headers)
    for row_obj in rows:
        ws.append(row_obj['cols'][:40])
    wb.save(out_path)


if __name__ == '__main__':
    app.run(debug=False, port=5050)
